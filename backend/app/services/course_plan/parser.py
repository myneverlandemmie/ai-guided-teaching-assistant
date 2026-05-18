"""课程计划 Excel 解析服务。

本模块只负责把固定格式的 `.xlsx` 授课计划解析为 planned lesson 字典列表。
不处理数据库保存、文件上传、登录认证或页面展示。
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - 运行环境缺依赖时给出明确错误
    load_workbook = None


REQUIRED_COLUMNS = ("week", "lesson_no", "hours", "content_raw")

COLUMN_ALIASES = {
    "week": ("周次", "*周次"),
    "lesson_no": ("课次", "*课次"),
    "hours": ("学时", "*学时"),
    "content_raw": ("教学内容(课堂教学，课带实验)", "教学内容", "*教学内容"),
    "tools": ("教学用具", "教学资源"),
    "homework": ("作业", "作业要求"),
    "notes": ("备注", "说明"),
}


class CoursePlanParseError(ValueError):
    """授课计划解析失败。

    用于向调用方返回明确、可展示的解析错误，例如缺少必要表头或文件为空。
    """


def normalize_cell_value(value: Any) -> str:
    """规范化 Excel 单元格值。

    Args:
        value: openpyxl 读取出的原始单元格值。

    Returns:
        去除首尾空白后的字符串；空值返回空字符串；整数型浮点数去掉 `.0`。

    Raises:
        不主动抛出业务异常。
    """

    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    """规范化表头，忽略空白与必填星号差异。"""

    text = normalize_cell_value(value)
    text = re.sub(r"\s+", "", text)
    return text.lstrip("*")


def _to_halfwidth_digits(value: str) -> str:
    """将全角数字转换为半角数字。"""

    return value.translate(str.maketrans("０１２３４５６７８９", "0123456789"))


def parse_lesson_code_and_title(content: Any) -> tuple[str, str]:
    """从教学内容中解析课次编码和课次标题。

    Args:
        content: 教学内容原文，例如 `0401-简单查询`。

    Returns:
        `(lesson_code, lesson_title)`。无法识别 4 位开头编码时，编码为空，标题使用教学内容。

    Raises:
        不主动抛出业务异常。
    """

    content_text = normalize_cell_value(content)
    if not content_text:
        return "", ""

    # 业务规则：教学内容以 4 位半角或全角数字开头时，将其作为课次编码。
    code_match = re.match(r"^([0-9０-９]{4})", content_text)
    if not code_match:
        return "", content_text

    lesson_code = _to_halfwidth_digits(code_match.group(1))
    title = content_text[code_match.end():].lstrip()
    if title[:1] in {"-", "—", "：", ":"}:
        title = title[1:].lstrip()

    return lesson_code, title.strip() or content_text


def detect_columns(headers: list[Any] | tuple[Any, ...]) -> dict[str, int]:
    """识别授课计划表头列。

    Args:
        headers: Excel 中某一行的表头单元格值列表。

    Returns:
        字段名到 0-based 列索引的映射，包含可识别的必需和可选字段。

    Raises:
        CoursePlanParseError: 缺少周次、课次、学时或教学内容等必要字段。
    """

    normalized_headers = [_normalize_header(header) for header in headers]
    detected: dict[str, int] = {}

    for field_name, aliases in COLUMN_ALIASES.items():
        normalized_aliases = {_normalize_header(alias) for alias in aliases}
        for index, header in enumerate(normalized_headers):
            if header in normalized_aliases:
                detected[field_name] = index
                break

    missing = [field for field in REQUIRED_COLUMNS if field not in detected]
    if missing:
        missing_names = ", ".join(missing)
        raise CoursePlanParseError(f"授课计划格式不符合当前版本要求，缺少必要字段: {missing_names}")

    return detected


def build_planned_lesson(row_data: dict[str, Any]) -> dict[str, str]:
    """将单行原始数据转换为 planned lesson 输出结构。

    Args:
        row_data: 已按字段名整理的行数据，至少包含 week、lesson_no、hours、content_raw。

    Returns:
        planned lesson 字典，包含本轮要求的 10 个字段。

    Raises:
        CoursePlanParseError: 行数据缺少有效行所需字段。
    """

    normalized = {key: normalize_cell_value(value) for key, value in row_data.items()}
    missing = [field for field in REQUIRED_COLUMNS if not normalized.get(field)]
    if missing:
        missing_names = ", ".join(missing)
        raise CoursePlanParseError(f"该行缺少有效课次所需字段: {missing_names}")

    lesson_code, lesson_title = parse_lesson_code_and_title(normalized["content_raw"])

    return {
        "week": normalized["week"],
        "lesson_no": normalized["lesson_no"],
        "hours": normalized["hours"],
        "lesson_code": lesson_code,
        "lesson_title": lesson_title,
        "content_raw": normalized["content_raw"],
        "tools": normalized.get("tools", ""),
        "homework": normalized.get("homework", ""),
        "notes": normalized.get("notes", ""),
        "status": "pending",
    }


def _find_header_row(worksheet: Any) -> tuple[int, dict[str, int]]:
    """在工作表中查找第一行可识别表头。"""

    last_error: CoursePlanParseError | None = None
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if not any(normalize_cell_value(cell) for cell in row):
            continue
        try:
            return row_number, detect_columns(row)
        except CoursePlanParseError as exc:
            last_error = exc
            continue

    if last_error is not None:
        raise last_error
    raise CoursePlanParseError("授课计划为空，未识别到表头字段。")


def parse_course_plan_xlsx(file_path: str | Path) -> list[dict[str, str]]:
    """解析 `.xlsx` 授课计划文件。

    Args:
        file_path: `.xlsx` 文件路径。

    Returns:
        planned lesson 字典列表。空行和缺少有效行必要字段的行会被跳过。

    Raises:
        CoursePlanParseError: 文件类型不支持、文件不存在、缺少必要表头或未解析到有效课次。
        ImportError: 当前 Python 环境未安装 openpyxl。
    """

    if load_workbook is None:
        raise ImportError("解析 .xlsx 授课计划需要安装 openpyxl。")

    path = Path(file_path)
    if path.suffix.lower() != ".xlsx":
        raise CoursePlanParseError("当前仅支持 .xlsx 格式的授课计划，请重新上传。")
    if not path.exists():
        raise CoursePlanParseError(f"授课计划文件不存在: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row_number, columns = _find_header_row(worksheet)

    planned_lessons: list[dict[str, str]] = []
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_number <= header_row_number:
            continue
        if not any(normalize_cell_value(cell) for cell in row):
            continue

        row_data = {
            field_name: row[column_index] if column_index < len(row) else ""
            for field_name, column_index in columns.items()
        }

        # 有效教学行至少要有周次、课次、学时、教学内容；不足则不进入 planned lessons。
        if any(not normalize_cell_value(row_data.get(field)) for field in REQUIRED_COLUMNS):
            continue

        planned_lessons.append(build_planned_lesson(row_data))

    if not planned_lessons:
        raise CoursePlanParseError("未识别到有效课次，请检查授课计划内容。")

    return planned_lessons
