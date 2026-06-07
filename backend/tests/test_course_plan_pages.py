from collections.abc import Generator
from pathlib import Path
import re

import httpx
import pytest
from openpyxl import load_workbook
from sqlalchemy import String, Text, create_engine, inspect as sa_inspect, select, text
from sqlalchemy.orm import Session, sessionmaker

from app import main
from app.db.base import create_database_tables
from app.models.course import Course
from app.models.course_plan import PlannedLesson
from app.models.knowledge_outline import KnowledgeOutline
from app.models.lesson import Lesson, LessonMaterial
from app.models.lesson_draft import LessonDraft
from app.services.ai.deepseek_client import (
    DeepSeekConfig,
    DeepSeekProviderError,
    build_knowledge_outline_prompt,
    generate_deepseek_knowledge_outline,
    get_allowed_deepseek_models,
    get_deepseek_config,
    get_default_deepseek_model,
    is_allowed_deepseek_model,
)
from app.services.ai.provider import GeneratedOutline
from app.services.ai.lesson_draft_service import build_chaoxing_catalog
from app.services.ai.sanitizer import sanitize_text_for_outline
from app.services.ai.session_key_store import clear_all_session_api_keys_for_tests


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SAMPLE_PLAN = PROJECT_ROOT / "data" / "sample-course-plans" / "2025-2026-database-course-plan.xlsx"
SAME_ORIGIN_HEADERS = {"origin": "http://testserver"}


@pytest.fixture
def anyio_backend() -> str:
    return "asyncio"


@pytest.fixture(autouse=True)
def inline_threadpool_for_tests(monkeypatch: pytest.MonkeyPatch) -> None:
    """测试环境不启动真实线程，避免误触外部请求或沙箱线程限制。"""

    async def run_inline(func: object, *args: object, **kwargs: object) -> object:
        return func(*args, **kwargs)  # type: ignore[misc]

    monkeypatch.setattr(main, "run_in_threadpool", run_inline)


def _build_test_client(tmp_path: Path) -> tuple[httpx.AsyncClient, sessionmaker[Session]]:
    clear_all_session_api_keys_for_tests()
    database_path = tmp_path / "test-course-plan-pages.sqlite"
    engine = create_engine(
        f"sqlite+pysqlite:///{database_path}",
        connect_args={"check_same_thread": False},
        future=True,
    )
    create_database_tables(engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    async def override_get_db() -> Generator[Session, None, None]:
        db = session_factory()
        try:
            yield db
        finally:
            db.close()

    main.app.dependency_overrides[main.get_db] = override_get_db
    main.COURSE_PLAN_UPLOAD_DIR = tmp_path / "course-plans"
    main.LESSON_MATERIAL_UPLOAD_DIR = tmp_path / "lesson-materials"
    main.CHAOXING_EXPORT_DIR = tmp_path / "exports" / "chaoxing"
    main.GUIDE_EXPORT_DIR = tmp_path / "exports" / "guides"
    transport = httpx.ASGITransport(app=main.app)
    return httpx.AsyncClient(transport=transport, base_url="http://testserver"), session_factory


def _create_course(session_factory: sessionmaker[Session]) -> Course:
    with session_factory() as session:
        course = Course(title="数据库应用与数据分析", semester="2025-2026-2", status="draft")
        session.add(course)
        session.commit()
        session.refresh(course)
        return course


def _database_contains_text(session: Session, needle: str) -> bool:
    """扫描业务表文本字段是否包含指定内容，不返回字段值。"""

    bind = session.get_bind()
    inspector = sa_inspect(bind)
    for table_name in inspector.get_table_names():
        if table_name.startswith("sqlite_"):
            continue
        columns = [
            column["name"]
            for column in inspector.get_columns(table_name)
            if isinstance(column["type"], (String, Text))
        ]
        if not columns:
            continue
        quoted_columns = ", ".join(f'"{column}"' for column in columns)
        rows = session.execute(text(f'SELECT {quoted_columns} FROM "{table_name}"')).all()
        for row in rows:
            if any(value is not None and needle in str(value) for value in row):
                return True
    return False


async def _upload_sample_plan(
    client: httpx.AsyncClient,
    course: Course,
) -> str:
    response = await client.post(
        f"/courses/{course.id}/course-plan/upload",
        files={
            "file": (
                SAMPLE_PLAN.name,
                SAMPLE_PLAN.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    return response.headers["location"]


async def _create_first_lesson(client: httpx.AsyncClient, session_factory: sessionmaker[Session], course: Course) -> None:
    """通过授课计划确认流程创建一个正式课次。"""

    await _upload_sample_plan(client, course)
    with session_factory() as session:
        selected_id = session.scalar(select(PlannedLesson.id).order_by(PlannedLesson.id))
        assert selected_id is not None
    response = await client.post(
        "/course-plan-uploads/1/confirm",
        data={"planned_lesson_ids": str(selected_id)},
        follow_redirects=False,
    )
    assert response.status_code == 303


def _create_reviewed_outline(session_factory: sessionmaker[Session], lesson_id: int = 1) -> None:
    """为导学草稿测试准备一条已复核知识主干。"""

    with session_factory() as session:
        outline = KnowledgeOutline(
            lesson_id=lesson_id,
            ai_raw_output="知识主干：核心概念、课堂任务、易错点和职业素养提示。",
            edited_content="知识主干：核心概念、课堂任务、易错点和职业素养提示。学生需要完成基础操作并复盘错误。",
            status="reviewed",
            generated_by_model="test-model",
        )
        session.add(outline)
        session.commit()


@pytest.mark.anyio
async def test_lesson_drafts_page_requires_knowledge_outline(tmp_path: Path) -> None:
    client, session_factory = _build_test_client(tmp_path)
    course = _create_course(session_factory)
    try:
        await _create_first_lesson(client, session_factory, course)

        response = await client.get("/lessons/1/drafts")

        assert response.status_code == 200
        assert "课前学情与学生导学案" in response.text
        assert "请先生成并保存知识主干" in response.text
        assert "本系统不做学生答题、不做发布、不做统计、不对接学习通 API" in response.text
        with session_factory() as session:
            assert session.scalars(select(LessonDraft)).all() == []
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()


@pytest.mark.anyio
async def test_default_lesson_draft_generation_creates_probe_and_low_guide(tmp_path: Path) -> None:
    client, session_factory = _build_test_client(tmp_path)
    course = _create_course(session_factory)
    try:
        await _create_first_lesson(client, session_factory, course)
        _create_reviewed_outline(session_factory)

        diagnostic_response = await client.post("/lessons/1/drafts/generate/diagnostic_probe", follow_redirects=False)
        response = await client.post("/lessons/1/drafts/generate/guide_low", follow_redirects=False)

        assert diagnostic_response.status_code == 303
        assert response.status_code == 303
        assert response.headers["location"] == "/lessons/1/drafts?draft_fallback=1"
        with session_factory() as session:
            drafts = session.scalars(select(LessonDraft).order_by(LessonDraft.draft_type)).all()
            assert len(drafts) == 2
            assert {draft.draft_type for draft in drafts} == {
                "diagnostic_probe",
                "guide_low",
            }
            assert {draft.status for draft in drafts} == {"draft"}
            assert {draft.generated_by for draft in drafts} == {"local-structured-draft"}

            diagnostic = next(draft for draft in drafts if draft.draft_type == "diagnostic_probe")
            for text_value in [
                "题目",
                "参考答案",
                "简短解析",
                "诊断点",
                "难度",
                "基础版主文档建议",
                "提升任务包建议",
                "拓展挑战包建议",
            ]:
                assert text_value in diagnostic.content
            assert "本前测用于判断学习起点，不作为正式考试成绩" in diagnostic.content

            guide_contents = {draft.draft_type: draft.content for draft in drafts}
            assert "基础版导学案" in guide_contents["guide_low"]
            four_char_headings = [
                "学习导航",
                "学习情境",
                "知识要点",
                "边学边填",
                "例题引路",
                "仿做练习",
                "过程记录",
                "重点速记",
                "带回小练",
                "学习记录",
                "学习自评",
            ]
            for content in [guide_contents["guide_low"]]:
                for heading in four_char_headings:
                    assert heading in content
                assert "AI 草稿声明" in content
                assert "rule-based" not in content
                assert "rule_based" not in content
                assert "mock" not in content

        page_response = await client.get("/lessons/1/drafts")
        assert page_response.status_code == 200
        assert "课前学情测试" in page_response.text
        assert "课前学情与学生导学案" in page_response.text
        assert "以下内容为教师草稿，仅供审阅、修改、复制与导出" in page_response.text
        assert "系统不会自动发布给学生" in page_response.text
        assert "全班通用导学案" in page_response.text
        assert "AI 生成工作台" in page_response.text
        assert "本页主文档" in page_response.text
        assert "巩固提升任务包" in page_response.text
        assert "拓展探究任务包" in page_response.text
        assert "生成课前学情测试" in page_response.text
        assert "生成基础版导学案" in page_response.text
        assert "生成巩固提升任务包" in page_response.text
        assert "生成拓展探究任务包" in page_response.text
        assert "AI 正在生成，请稍候" in page_response.text
        assert "请勿重复点击" in page_response.text
        assert "本地结构化草稿" in page_response.text
        assert "rule_based" not in page_response.text
        assert "rule-based" not in page_response.text
        assert "mock" not in page_response.text
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()


@pytest.mark.anyio
async def test_can_generate_mid_and_high_guides_after_low_guide_exists(tmp_path: Path) -> None:
    client, session_factory = _build_test_client(tmp_path)
    course = _create_course(session_factory)
    try:
        await _create_first_lesson(client, session_factory, course)
        _create_reviewed_outline(session_factory)
        await client.post("/lessons/1/drafts/generate/diagnostic_probe", follow_redirects=False)
        await client.post("/lessons/1/drafts/generate/guide_low", follow_redirects=False)

        mid_response = await client.post("/lessons/1/drafts/generate/guide_mid", follow_redirects=False)
        high_response = await client.post("/lessons/1/drafts/generate/guide_high", follow_redirects=False)

        assert mid_response.status_code == 303
        assert high_response.status_code == 303
        with session_factory() as session:
            drafts = session.scalars(select(LessonDraft).order_by(LessonDraft.draft_type)).all()
            assert len(drafts) == 4
            draft_types = {draft.draft_type for draft in drafts}
            assert draft_types == {"diagnostic_probe", "guide_low", "guide_mid", "guide_high"}
            guide_mid = next(draft for draft in drafts if draft.draft_type == "guide_mid")
            guide_high = next(draft for draft in drafts if draft.draft_type == "guide_high")
            assert "提升任务包" in guide_mid.content
            assert "拓展挑战包" in guide_high.content
            assert "学习导航" not in guide_mid.content
            assert "学习导航" not in guide_high.content
            assert "不是完整导学案" in guide_mid.content
            assert "不是完整导学案" in guide_high.content

            for draft in [guide_mid, guide_high]:
                download_response = await client.get(f"/lessons/1/drafts/{draft.id}/download-md")
                assert download_response.status_code == 200
                assert "text/markdown" in download_response.headers["content-type"]
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()


@pytest.mark.anyio
async def test_diagnostic_probe_exports_chaoxing_template(tmp_path: Path) -> None:
    client, session_factory = _build_test_client(tmp_path)
    course = _create_course(session_factory)
    try:
        await _create_first_lesson(client, session_factory, course)
        _create_reviewed_outline(session_factory)
        await client.post("/lessons/1/drafts/generate/diagnostic_probe", follow_redirects=False)
        with session_factory() as session:
            saved_course = session.get(Course, course.id)
            lesson = session.get(Lesson, 1)
            assert saved_course is not None
            assert lesson is not None
            saved_course.title = "数据库基础"
            lesson.lesson_code = "0406"
            lesson.title = "分组查询"
            session.commit()
            draft = session.scalar(select(LessonDraft).where(LessonDraft.draft_type == "diagnostic_probe"))
            assert draft is not None
            draft_id = draft.id

        export_response = await client.post(
            f"/lessons/1/drafts/{draft_id}/export-chaoxing",
            follow_redirects=False,
        )

        assert export_response.status_code == 303
        assert "chaoxing_file=" in export_response.headers["location"]
        export_files = list((tmp_path / "exports" / "chaoxing").glob("*.xlsx"))
        assert len(export_files) == 1
        workbook = load_workbook(export_files[0])
        worksheet = workbook["课程题库"]
        headers = [cell.value for cell in worksheet[1]]
        for header in ["目录", "题目类型", "大题题干", "正确答案", "答案解析", "难易度", "知识点", "标签", "选项数", "选项A", "选项B"]:
            assert header in headers
        rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 5
        catalogs = {row[0] for row in rows}
        assert "/数据库基础/0406-分组查询" in catalogs
        assert all("智学导评" not in str(catalog) for catalog in catalogs)
        question_types = {row[1] for row in rows}
        assert {"单选题", "判断题", "填空题"}.issubset(question_types)
        judgment_row = next(row for row in rows if row[1] == "判断题")
        assert judgment_row[10] == 2
        assert judgment_row[11] == "正确"
        assert judgment_row[12] == "错误"

        page_response = await client.get(export_response.headers["location"])
        assert page_response.status_code == 200
        assert "下载学习通题库模板" in page_response.text
        assert "/exports/chaoxing/" in page_response.text
        assert "学习通 API" in page_response.text
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()


def test_chaoxing_catalog_falls_back_without_course_name() -> None:
    lesson = Lesson(
        id=1,
        course_id=1,
        planned_lesson_id=None,
        week="1",
        lesson_no="1",
        hours="2",
        lesson_code="0406",
        title="分组查询",
        content_summary="分组查询",
        status="draft",
    )

    assert build_chaoxing_catalog(lesson) == "/0406-分组查询"


@pytest.mark.anyio
async def test_guide_low_can_download_markdown(tmp_path: Path) -> None:
    client, session_factory = _build_test_client(tmp_path)
    course = _create_course(session_factory)
    try:
        await _create_first_lesson(client, session_factory, course)
        _create_reviewed_outline(session_factory)
        await client.post("/lessons/1/drafts/generate/guide_low", follow_redirects=False)
        with session_factory() as session:
            draft = session.scalar(select(LessonDraft).where(LessonDraft.draft_type == "guide_low"))
            assert draft is not None
            draft_id = draft.id
            expected_content = draft.content

        response = await client.get(f"/lessons/1/drafts/{draft_id}/download-md")

        assert response.status_code == 200
        assert "text/markdown" in response.headers["content-type"]
        assert response.text == expected_content
        assert "学习导航" in response.text
        assert "rule-based" not in response.text
        assert "rule_based" not in response.text
        assert "mock" not in response.text
        markdown_file = tmp_path / "exports" / "guides" / "lesson_1_core_learning_guide.md"
        assert markdown_file.exists()
        assert markdown_file.read_text(encoding="utf-8") == expected_content
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()


@pytest.mark.anyio
async def test_lesson_draft_can_be_edited_and_saved(tmp_path: Path) -> None:
    client, session_factory = _build_test_client(tmp_path)
    course = _create_course(session_factory)
    try:
        await _create_first_lesson(client, session_factory, course)
        _create_reviewed_outline(session_factory)
        await client.post("/lessons/1/drafts/generate/guide_low", follow_redirects=False)
        await client.post("/lessons/1/drafts/generate/guide_mid", follow_redirects=False)
        with session_factory() as session:
            draft = session.scalar(select(LessonDraft).where(LessonDraft.draft_type == "guide_mid"))
            assert draft is not None
            draft_id = draft.id

        response = await client.post(
            f"/lessons/1/drafts/{draft_id}/save",
            data={"title": "教师修改后的提升任务包", "content": "教师已修改：保留关键提示并增加半开放任务。"},
            follow_redirects=False,
        )

        assert response.status_code == 303
        assert response.headers["location"] == "/lessons/1/drafts"
        with session_factory() as session:
            saved = session.get(LessonDraft, draft_id)
            assert saved is not None
            assert saved.title == "教师修改后的提升任务包"
            assert saved.content == "教师已修改：保留关键提示并增加半开放任务。"
            assert saved.status == "reviewed"
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()


@pytest.mark.anyio
async def test_regenerating_lesson_drafts_upserts_current_drafts(tmp_path: Path) -> None:
    client, session_factory = _build_test_client(tmp_path)
    course = _create_course(session_factory)
    try:
        await _create_first_lesson(client, session_factory, course)
        _create_reviewed_outline(session_factory)

        first_response = await client.post("/lessons/1/drafts/generate/diagnostic_probe", follow_redirects=False)
        second_response = await client.post("/lessons/1/drafts/generate/diagnostic_probe", follow_redirects=False)

        assert first_response.status_code == 303
        assert second_response.status_code == 303
        with session_factory() as session:
            drafts = session.scalars(select(LessonDraft)).all()
            assert len(drafts) == 1
            assert len({(draft.lesson_id, draft.draft_type) for draft in drafts}) == 1
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()


@pytest.mark.anyio
async def test_no_sql_python_c_grading_demo_routes_added(tmp_path: Path) -> None:
    client, _ = _build_test_client(tmp_path)
    try:
        for path in ["/demo-grading/sql", "/demo-grading/python", "/demo-grading/c"]:
            response = await client.get(path)
            assert response.status_code == 404
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()
