from collections.abc import Generator
from pathlib import Path

import httpx
import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

from app import main
from app.db.base import create_database_tables
from app.models.course import Course
from app.models.lesson import Lesson, LessonMaterial
from app.services.lesson_materials.document_text_extractor import EMPTY_XLSX_TEXT_MESSAGE


@pytest.fixture
def anyio_backend() -> str:
    return "asyncio"


def _build_test_client(tmp_path: Path) -> tuple[httpx.AsyncClient, sessionmaker[Session]]:
    database_path = tmp_path / "test-lesson-materials-xlsx.sqlite"
    engine = create_engine(
        f"sqlite+pysqlite:///{database_path}",
        connect_args={"check_same_thread": False},
        future=True,
    )
    create_database_tables(engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    async def override_get_db() -> Generator[Session, None, None]:
        db = session_factory()
        try:
            yield db
        finally:
            db.close()

    main.app.dependency_overrides[main.get_db] = override_get_db
    main.LESSON_MATERIAL_UPLOAD_DIR = tmp_path / "lesson-materials"
    transport = httpx.ASGITransport(app=main.app)
    return httpx.AsyncClient(transport=transport, base_url="http://testserver"), session_factory


def _create_lesson(session_factory: sessionmaker[Session]) -> int:
    with session_factory() as session:
        course = Course(title="测试课程", semester="2025-2026-2", status="draft")
        session.add(course)
        session.flush()
        lesson = Lesson(
            course_id=course.id,
            week="1",
            lesson_no="1",
            hours="2",
            lesson_code="0101",
            title="数据记录表分析",
            content_summary="读取表格数据并整理课堂记录。",
            homework_hint="整理记录表。",
            status="draft",
        )
        session.add(lesson)
        session.commit()
        return lesson.id


def _create_sample_xlsx(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "成绩记录"
    sheet.append(["姓名", "成绩", "备注"])
    sheet.append(["张三", 88, "已提交"])
    sheet.append([None, None, None])
    sheet.append(["李四", 76, "需订正"])
    steps = workbook.create_sheet("实验步骤")
    steps.append(["步骤", "说明"])
    steps.append(["1", "连接传感器"])
    steps.append(["2", "记录输出信号"])
    workbook.save(path)


def _create_empty_xlsx(path: Path) -> None:
    workbook = Workbook()
    workbook.save(path)


def _create_whitespace_xlsx(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "空白内容"
    sheet["A1"] = "   "
    sheet["B2"] = "\n\t  "
    workbook.save(path)


@pytest.mark.anyio
async def test_can_upload_xlsx_lesson_material_and_extract_sheet_text(tmp_path: Path) -> None:
    client, session_factory = _build_test_client(tmp_path)
    lesson_id = _create_lesson(session_factory)
    xlsx_path = tmp_path / "records.xlsx"
    _create_sample_xlsx(xlsx_path)
    try:
        response = await client.post(
            f"/lessons/{lesson_id}/materials",
            data={"title": "", "material_type": "evaluation_sheet", "content": ""},
            files={
                "files": (
                    "records.xlsx",
                    xlsx_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            follow_redirects=False,
        )

        assert response.status_code == 303
        with session_factory() as session:
            material = session.scalar(select(LessonMaterial))
            assert material is not None
            assert material.material_type == "evaluation_sheet"
            assert material.title == "0101-评价表"
            assert material.file_path is not None
            assert "# XLSX 表格资料提取" in material.content
            assert "## Sheet: 成绩记录" in material.content
            assert "## Sheet: 实验步骤" in material.content
            assert "姓名 成绩 备注" in material.content
            assert "张三 88 已提交" in material.content
            assert "李四 76 需订正" in material.content
            assert "连接传感器" in material.content
            assert EMPTY_XLSX_TEXT_MESSAGE not in material.content

        page_response = await client.get(f"/ui-v2/lessons/{lesson_id}/materials-outline")
        assert page_response.status_code == 200
        assert "资料类别：评价表 / 记录表" in page_response.text
        assert "records.xlsx" in page_response.text
        assert EMPTY_XLSX_TEXT_MESSAGE not in page_response.text
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()


@pytest.mark.anyio
async def test_empty_xlsx_lesson_material_upload_shows_friendly_message(tmp_path: Path) -> None:
    client, session_factory = _build_test_client(tmp_path)
    lesson_id = _create_lesson(session_factory)
    xlsx_path = tmp_path / "empty.xlsx"
    _create_empty_xlsx(xlsx_path)
    try:
        response = await client.post(
            f"/lessons/{lesson_id}/materials",
            data={"title": "空表", "material_type": "evaluation_sheet", "content": ""},
            files={
                "files": (
                    "empty.xlsx",
                    xlsx_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 400
        assert response.status_code != 500
        assert f"empty.xlsx：{EMPTY_XLSX_TEXT_MESSAGE}" in response.text
        assert str(tmp_path) not in response.text
        with session_factory() as session:
            assert session.scalars(select(LessonMaterial)).all() == []
        upload_dir = tmp_path / "lesson-materials"
        assert not upload_dir.exists() or list(upload_dir.iterdir()) == []
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()


@pytest.mark.anyio
async def test_whitespace_xlsx_lesson_material_error_stays_on_v2_page(tmp_path: Path) -> None:
    client, session_factory = _build_test_client(tmp_path)
    lesson_id = _create_lesson(session_factory)
    xlsx_path = tmp_path / "blank-cells.xlsx"
    _create_whitespace_xlsx(xlsx_path)
    return_to = f"/ui-v2/lessons/{lesson_id}/materials-outline"
    try:
        response = await client.post(
            f"/lessons/{lesson_id}/materials",
            data={
                "title": "空白表格",
                "input_mode": "file_upload",
                "material_category": "evaluation_sheet",
                "return_to": return_to,
            },
            files={
                "files": (
                    "blank-cells.xlsx",
                    xlsx_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 400
        assert response.status_code != 500
        assert "lesson-materials-v2" in response.text
        assert "课程资料整理" in response.text
        assert "测试课程" in response.text
        assert "数据记录表分析" in response.text
        assert f"blank-cells.xlsx：{EMPTY_XLSX_TEXT_MESSAGE}" in response.text
        assert "课次详情" not in response.text
        with session_factory() as session:
            assert session.scalars(select(LessonMaterial)).all() == []
        upload_dir = tmp_path / "lesson-materials"
        assert not upload_dir.exists() or list(upload_dir.iterdir()) == []
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()


@pytest.mark.anyio
async def test_xls_lesson_material_upload_is_rejected_with_friendly_message(tmp_path: Path) -> None:
    client, session_factory = _build_test_client(tmp_path)
    lesson_id = _create_lesson(session_factory)
    try:
        response = await client.post(
            f"/lessons/{lesson_id}/materials",
            data={"title": "旧版表格", "material_type": "evaluation_sheet", "content": ""},
            files={"files": ("legacy.xls", b"fake xls", "application/vnd.ms-excel")},
        )

        assert response.status_code == 400
        assert "暂不支持旧版 .xls 表格文件" in response.text
        assert "另存为 .xlsx" in response.text
        with session_factory() as session:
            assert session.scalars(select(LessonMaterial)).all() == []
    finally:
        await client.aclose()
        main.app.dependency_overrides.clear()
