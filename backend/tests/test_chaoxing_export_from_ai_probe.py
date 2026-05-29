from pathlib import Path

from openpyxl import load_workbook

from app.models.course import Course
from app.models.lesson import Lesson
from app.models.lesson_draft import LessonDraft
from app.services.ai.lesson_draft_service import (
    build_chaoxing_rows,
    parse_diagnostic_probe_questions,
    write_chaoxing_template_xlsx,
)


def _lesson() -> Lesson:
    course = Course(id=1, title="数据库基础", semester="2025-2026", status="active")
    lesson = Lesson(
        id=1,
        course_id=1,
        week="1",
        lesson_no="1",
        hours="2",
        lesson_code="0406",
        title="分组查询",
        content_summary="GROUP BY 分组查询。",
        status="draft",
    )
    lesson.course = course
    return lesson


def _draft(content: str) -> LessonDraft:
    return LessonDraft(
        id=1,
        lesson_id=1,
        source_outline_id=1,
        draft_type="diagnostic_probe",
        title="课前学情测试",
        content=content,
        status="draft",
        generated_by="deepseek-v4-flash",
    )


FLASH_LIKE_PROBE = """# 课前学情测试草稿

## 题目 1：分组查询前最需要确认什么？
题型：单选题
选项：A. 分组字段；B. 字体颜色；C. 座位；D. 无关内容
答案：A
解析：分组字段决定统计口径。
知识点：分组字段识别
难度：基础

### 题目 2
题型：判断题
题干：HAVING 通常用于对分组后的结果进行筛选。
正确答案：正确
答案解析：WHERE 和 HAVING 的筛选阶段不同。
对应知识点：HAVING 用法
难度：中等

**题目 3**
题型：填空题
题干：GROUP BY 按一个或多个字段的______进行分组。
参考答案：组合
简短解析：多字段分组可以理解为按字段组合进行分组。
诊断点：多字段分组
难度：基础
"""


PRO_LIKE_PROBE = """# 课前学情测试草稿

### 题目 1
- 题型：单选题
- 题干：分组统计时应优先确认哪一项？
- 选项：A. 统计口径；B. 字体；C. 座位；D. 无关内容
- 参考答案：A
- 简短解析：统计口径影响结果准确性。
- 诊断点：结果核验
- 难度：基础
"""


LOCAL_PROBE = """# 本地结构化前测

### 题目 1
- 题型：判断题
- 题干：课前学情测试不作为正式考试成绩。
- 参考答案：正确
- 简短解析：它用于学习起点诊断。
- 诊断点：前测定位
- 难度：基础
"""


def test_parse_flash_like_ai_probe_questions() -> None:
    questions = parse_diagnostic_probe_questions(FLASH_LIKE_PROBE)

    assert len(questions) == 3
    assert questions[0].prompt == "分组查询前最需要确认什么？"
    assert questions[0].question_type == "单选题"
    assert questions[0].answer == "A"
    assert questions[1].question_type == "判断题"
    assert questions[1].answer == "正确"
    assert questions[2].question_type == "填空题"
    assert questions[2].diagnosis_point == "多字段分组"


def test_ai_probe_rows_are_not_empty_for_flash_pro_and_local_outputs(tmp_path: Path) -> None:
    lesson = _lesson()
    for content in [FLASH_LIKE_PROBE, PRO_LIKE_PROBE, LOCAL_PROBE]:
        rows = build_chaoxing_rows(lesson, _draft(content))
        assert rows
        assert all(row[0] == "/数据库基础/0406-分组查询" for row in rows)
        assert all(row[2] for row in rows)
        assert all(row[5] for row in rows)
        assert all(row[6] for row in rows)

    output_path = tmp_path / "lesson_1_0406_diagnostic_probe.xlsx"
    write_chaoxing_template_xlsx(lesson, _draft(FLASH_LIKE_PROBE), output_path)
    workbook = load_workbook(output_path)
    worksheet = workbook["课程题库"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 3
    assert rows[1][1] == "判断题"
    assert rows[1][10] == 2
    assert rows[1][11] == "正确"
    assert rows[1][12] == "错误"
